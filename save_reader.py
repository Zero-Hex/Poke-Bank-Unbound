#!/usr/bin/env python3
"""
Pokemon Unbound Save Reader
Reads a .sav file and outputs party + PC boxes to an Excel spreadsheet.

Usage:
    python unbound_save_reader.py <save_file.sav> [output.xlsx]

Works with Pokemon Unbound and other CFRU-based FireRed ROM hacks.
Data files (pokemon.txt, abilities.txt, etc.) should be in a 'data/' subfolder,
copied from the PUSE backend/data/ directory.
"""

import struct
import sys
import os
import json
import math
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants (from PUSE)
# ---------------------------------------------------------------------------
SECTION_SIZE = 0x1000
CHECKSUM_LENGTH = 0xFF4
FOOTER_ID_OFF = 0xFF4       # Actually 0xFF4 for section ID in PUSE
FOOTER_CHK_OFF = 0xFF6
FOOTER_SAVEIDX_OFF = 0xFFC
TRAINER_SECTION_ID = 1

PARTY_OFFSET = 0x38
PARTY_COUNT_OFFSET = 0x34
MON_SIZE_PARTY = 100

# PC sectors (PUSE: sectors 5-12 are Pokemon stream)
POKEMON_STREAM_SECTORS = [5, 6, 7, 8, 9, 10, 11, 12]
PRESET_SECTOR_ID = 0
ALL_PC_SECTORS = POKEMON_STREAM_SECTORS + [13, PRESET_SECTOR_ID]
SECTOR_HEADER_SIZE = 4
SECTOR_PAYLOAD_SIZE = 0xFF0
MON_SIZE_PC = 58
BOX_SLOT_COUNT = 30
OFFSET_PRESET_START = 0xB0
PRESET_CAPACITY = 30

# Party Pokemon offsets (within 100-byte structure)
OFF_PARTY_PID       = 0x00
OFF_PARTY_OTID      = 0x04
OFF_PARTY_NICK      = 0x08
OFF_PARTY_DATA      = 0x20
OFF_PARTY_CHECKSUM  = 0x1C
OFF_PARTY_LEVEL     = 0x54
OFF_PARTY_CURR_HP   = 0x56
OFF_PARTY_MAX_HP    = 0x58
OFF_PARTY_ATK       = 0x5A
OFF_PARTY_DEF       = 0x5C
OFF_PARTY_SPE       = 0x5E
OFF_PARTY_SPA       = 0x60
OFF_PARTY_SPD       = 0x62

# PC Pokemon offsets (within 58-byte CFRU compact structure)
OFF_PC_PID      = 0x00
OFF_PC_OTID     = 0x04
OFF_PC_NICK     = 0x08
OFF_PC_SPECIES  = 0x1C
OFF_PC_ITEM     = 0x1E
OFF_PC_EXP      = 0x20
OFF_PC_MOVES    = 0x24   # pp_ups byte at 0x24, moves bitpacked at 0x27
OFF_PC_EVS      = 0x2C
OFF_PC_IVS      = 0x36

# Party Pokemon offsets — same compact structure but OT_NAME field is 4 bytes longer,
# shifting everything after the nickname by +4 compared to PC offsets.
OFF_PTY_PID     = 0x00
OFF_PTY_OTID    = 0x04
OFF_PTY_NICK    = 0x08
OFF_PTY_SPECIES = 0x20
OFF_PTY_ITEM    = 0x22
OFF_PTY_EXP     = 0x24
OFF_PTY_PP_UPS  = 0x28   # pp_ups byte
OFF_PTY_MOVES   = 0x2C   # moves as 4 plain u16s (NOT bitpacked like PC)
OFF_PTY_EVS     = 0x30   # EVs (plain bytes, confirmed)
# Note: IVs stay at 0x3A (confirmed)
OFF_PTY_IVS     = 0x3A
# Party-only stat bytes (after the compact block)
OFF_PTY_LEVEL   = 0x54
OFF_PTY_MAX_HP  = 0x58
OFF_PTY_ATK     = 0x5A
OFF_PTY_DEF     = 0x5C
OFF_PTY_SPE     = 0x5E
OFF_PTY_SPA     = 0x60
OFF_PTY_SPD     = 0x62

SHINY_THRESHOLD = 16

DB_NATURES = {
    0: "Hardy",   1: "Lonely",  2: "Brave",   3: "Adamant", 4: "Naughty",
    5: "Bold",    6: "Docile",  7: "Relaxed", 8: "Impish",  9: "Lax",
   10: "Timid",  11: "Hasty",  12: "Serious",13: "Jolly",  14: "Naive",
   15: "Modest", 16: "Mild",   17: "Quiet",  18: "Bashful",19: "Rash",
   20: "Calm",   21: "Gentle", 22: "Sassy",  23: "Careful",24: "Quirky"
}

NATURE_EFFECTS = {
    # nature_id: (boosted_stat, lowered_stat)  None = neutral
    0: (None, None),       # Hardy
    1: ("Atk", "Def"),     # Lonely
    2: ("Atk", "Spe"),     # Brave
    3: ("Atk", "SpA"),     # Adamant
    4: ("Atk", "SpD"),     # Naughty
    5: ("Def", "Atk"),     # Bold
    6: (None, None),       # Docile
    7: ("Def", "Spe"),     # Relaxed
    8: ("Def", "SpA"),     # Impish
    9: ("Def", "SpD"),     # Lax
   10: ("Spe", "Atk"),     # Timid
   11: ("Spe", "Def"),     # Hasty
   12: (None, None),       # Serious
   13: ("Spe", "SpA"),     # Jolly
   14: ("Spe", "SpD"),     # Naive
   15: ("SpA", "Atk"),     # Modest
   16: ("SpA", "Def"),     # Mild
   17: ("SpA", "Spe"),     # Quiet
   18: (None, None),       # Bashful
   19: ("SpA", "SpD"),     # Rash
   20: ("SpD", "Atk"),     # Calm
   21: ("SpD", "Def"),     # Gentle
   22: ("SpD", "Spe"),     # Sassy
   23: ("SpD", "SpA"),     # Careful
   24: (None, None),       # Quirky
}

# PC charmap (full Unbound/CFRU set from PUSE pc.py)
CHARMAP_PC = {
    0x00: " ", 0x01: "À", 0x02: "Á", 0x03: "Â", 0x04: "Ç", 0x05: "È",
    0x06: "É", 0x07: "Ê", 0x08: "Ë", 0x09: "Ì", 0x0B: "Î", 0x0C: "Ï",
    0x0D: "Ò", 0x0E: "Ó", 0x0F: "Ô", 0x10: "Œ", 0x11: "Ù", 0x12: "Ú",
    0x13: "Û", 0x14: "Ñ", 0x15: "ß", 0x16: "à", 0x17: "á", 0x19: "ç",
    0x1A: "è", 0x1B: "é", 0x1C: "ê", 0x1D: "ë", 0x1E: "ì", 0x20: "î",
    0x21: "ï", 0x22: "ò", 0x23: "ó", 0x24: "ô", 0x25: "œ", 0x26: "ù",
    0x27: "ú", 0x28: "û", 0x29: "ñ", 0x2A: "º", 0x2B: "ª", 0x2D: "&",
    0x2E: "+", 0x34: "Lv", 0x35: "=", 0x36: ";", 0x51: "¿", 0x52: "¡",
    0x53: "PK", 0x54: "MN", 0x55: "PO", 0x56: "Ké", 0x57: "Bl",
    0x58: "oc", 0x59: "k", 0x5A: "Í", 0x5B: "%", 0x5C: "(", 0x5D: ")",
    0x68: "â", 0x6F: "í", 0x79: "↑", 0x7A: "↓", 0x7B: "←", 0x7C: "→",
    0x85: "<", 0x86: ">", 0xA1: "0", 0xA2: "1", 0xA3: "2", 0xA4: "3",
    0xA5: "4", 0xA6: "5", 0xA7: "6", 0xA8: "7", 0xA9: "8", 0xAA: "9",
    0xAB: "!", 0xAC: "?", 0xAD: ".", 0xAE: "-", 0xAF: "·", 0xB0: "...",
    0xB1: "\u201c", 0xB2: "\u201d", 0xB3: "\u2018", 0xB4: "\u2019",
    0xB5: "♂", 0xB6: "♀", 0xB7: "$", 0xB8: ",", 0xB9: "×", 0xBA: "/",
    0xBB: "A", 0xBC: "B", 0xBD: "C", 0xBE: "D", 0xBF: "E", 0xC0: "F",
    0xC1: "G", 0xC2: "H", 0xC3: "I", 0xC4: "J", 0xC5: "K", 0xC6: "L",
    0xC7: "M", 0xC8: "N", 0xC9: "O", 0xCA: "P", 0xCB: "Q", 0xCC: "R",
    0xCD: "S", 0xCE: "T", 0xCF: "U", 0xD0: "V", 0xD1: "W", 0xD2: "X",
    0xD3: "Y", 0xD4: "Z", 0xD5: "a", 0xD6: "b", 0xD7: "c", 0xD8: "d",
    0xD9: "e", 0xDA: "f", 0xDB: "g", 0xDC: "h", 0xDD: "i", 0xDE: "j",
    0xDF: "k", 0xE0: "l", 0xE1: "m", 0xE2: "n", 0xE3: "o", 0xE4: "p",
    0xE5: "q", 0xE6: "r", 0xE7: "s", 0xE8: "t", 0xE9: "u", 0xEA: "v",
    0xEB: "w", 0xEC: "x", 0xED: "y", 0xEE: "z", 0xEF: "▶", 0xF0: ":",
    0xF1: "Ä", 0xF2: "Ö", 0xF3: "Ü", 0xF4: "ä", 0xF5: "ö", 0xF6: "ü",
    0xFF: ""
}

# Party charmap (simpler, from PUSE party.py)
CHARMAP_PARTY = {
    0x00: " ", 0xAB: "!", 0xAC: "?", 0xAD: ".", 0xAE: "-", 0xFF: "",
    0xB0: "0", 0xB1: "1", 0xB2: "2", 0xB3: "3", 0xB4: "4",
    0xB5: "5", 0xB6: "6", 0xB7: "7", 0xB8: "8", 0xB9: "9",
    0xBB: "A", 0xBC: "B", 0xBD: "C", 0xBE: "D", 0xBF: "E", 0xC0: "F",
    0xC1: "G", 0xC2: "H", 0xC3: "I", 0xC4: "J", 0xC5: "K", 0xC6: "L",
    0xC7: "M", 0xC8: "N", 0xC9: "O", 0xCA: "P", 0xCB: "Q", 0xCC: "R",
    0xCD: "S", 0xCE: "T", 0xCF: "U", 0xD0: "V", 0xD1: "W", 0xD2: "X",
    0xD3: "Y", 0xD4: "Z",
    0xD5: "a", 0xD6: "b", 0xD7: "c", 0xD8: "d", 0xD9: "e", 0xDA: "f",
    0xDB: "g", 0xDC: "h", 0xDD: "i", 0xDE: "j", 0xDF: "k", 0xE0: "l",
    0xE1: "m", 0xE2: "n", 0xE3: "o", 0xE4: "p", 0xE5: "q", 0xE6: "r",
    0xE7: "s", 0xE8: "t", 0xE9: "u", 0xEA: "v", 0xEB: "w", 0xEC: "x",
    0xED: "y", 0xEE: "z",
}

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def find_data_dir():
    """Look for the data directory relative to this script."""
    candidates = [
        Path(__file__).parent / "data",
        Path.cwd() / "data",
        Path(__file__).parent / "backend" / "data",
        Path.cwd() / "backend" / "data",
    ]
    for p in candidates:
        if (p / "pokemon.txt").exists():
            return p
    return None

def load_id_name_file(path):
    out = {}
    if not path or not Path(path).exists():
        return out
    for line in Path(path).read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or ":" not in line:
            continue
        left, right = line.split(":", 1)
        if left.strip().isdigit():
            out[int(left.strip())] = right.strip()
    return out

def load_json_file(path):
    if not path or not Path(path).exists():
        return {}
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return {}

def load_databases(data_dir):
    if data_dir is None:
        print("WARNING: No 'data/' directory found. Species/move names will show as IDs.")
        print("         Copy the 'data/' folder from PUSE's backend/ directory next to this script.")
        return {}, {}, {}, {}, {}, {}, {}, {}

    db_species  = load_id_name_file(data_dir / "pokemon.txt")
    db_items    = load_id_name_file(data_dir / "items.txt")
    db_moves    = load_id_name_file(data_dir / "moves.txt")
    db_abilities = load_id_name_file(data_dir / "abilities.txt")

    raw_abilities_meta = load_json_file(data_dir / "species_abilities_meta.json")
    db_abilities_meta = {}
    for sid, meta in raw_abilities_meta.items():
        if str(sid).isdigit() and isinstance(meta, dict):
            db_abilities_meta[int(sid)] = {
                "a1": meta.get("ability_1_id"),
                "a2": meta.get("ability_2_id"),
                "ha": meta.get("hidden_ability_id"),
            }

    raw_identity = load_json_file(data_dir / "species_identity_meta.json")
    db_gender = {}
    for sid, meta in raw_identity.items():
        if str(sid).isdigit() and isinstance(meta, dict):
            t = meta.get("gender_threshold")
            if t is not None:
                db_gender[int(sid)] = int(t) & 0xFF

    raw_growth = load_json_file(data_dir / "species_growth_rates.json")
    db_growth_rates = {}
    for sid, meta in raw_growth.items():
        if str(sid).isdigit() and isinstance(meta, dict):
            rate = meta.get("growth_rate")
            if rate is not None:
                rate_int = int(rate)
                if 0 <= rate_int <= 5:
                    db_growth_rates[int(sid)] = rate_int

    raw_moves_json = load_json_file(data_dir / "move_table_from_rom.json")
    db_move_pp = {}
    rows = raw_moves_json.get("moves") if isinstance(raw_moves_json, dict) else None
    if isinstance(rows, list):
        for row in rows:
            try:
                db_move_pp[int(row["move_id"])] = int(row["base_pp"])
            except (KeyError, TypeError, ValueError):
                pass

    print(f"Loaded: {len(db_species)} species, {len(db_items)} items, "
          f"{len(db_moves)} moves, {len(db_abilities)} abilities, "
          f"{len(db_growth_rates)} growth rates")
    return db_species, db_items, db_moves, db_abilities, db_abilities_meta, db_gender, db_move_pp, db_growth_rates

# ---------------------------------------------------------------------------
# Binary helpers
# ---------------------------------------------------------------------------

def ru8(b, o):  return struct.unpack_from("<B", b, o)[0]
def ru16(b, o): return struct.unpack_from("<H", b, o)[0]
def ru32(b, o): return struct.unpack_from("<I", b, o)[0]

def decode_text(data, charmap):
    s = ""
    for byte in data:
        if byte == 0xFF:
            break
        s += charmap.get(byte, "?")
    return s.strip()

def is_shiny(otid, pid):
    tid = otid & 0xFFFF
    sid = (otid >> 16) & 0xFFFF
    sv = tid ^ sid ^ (pid & 0xFFFF) ^ ((pid >> 16) & 0xFFFF)
    return sv < SHINY_THRESHOLD

def gender_from_pid(pid, threshold):
    if threshold is None:        return "Unknown"
    if threshold == 255:         return "Genderless"
    if threshold == 0:           return "Male"
    if threshold == 254:         return "Female"
    return "Female" if (pid & 0xFF) < threshold else "Male"

def unpack_ivs(packed):
    return {
        "HP":  (packed >>  0) & 0x1F,
        "Atk": (packed >>  5) & 0x1F,
        "Def": (packed >> 10) & 0x1F,
        "Spe": (packed >> 15) & 0x1F,
        "SpA": (packed >> 20) & 0x1F,
        "SpD": (packed >> 25) & 0x1F,
    }

def exp_at_level(rate, n):
    """Return the EXP threshold for the start of level n (from PUSE party.py)."""
    if n <= 1: return 0
    if n > 100: n = 100
    if rate == 0: return n ** 3
    if rate == 1:
        if n <= 50:  return int((n**3 * (100 - n)) / 50)
        if n <= 68:  return int((n**3 * (150 - n)) / 100)
        if n <= 98:  return int((n**3 * ((1911 - 10*n) / 3)) / 500)
        return int((n**3 * (160 - n)) / 100)
    if rate == 2:
        if n <= 15:  return int(n**3 * ((math.floor((n+1)/3) + 24) / 50))
        if n <= 36:  return int(n**3 * ((n + 14) / 50))
        return int(n**3 * ((math.floor(n/2) + 32) / 50))
    if rate == 3: return int(1.2*(n**3) - 15*(n**2) + 100*n - 140)
    if rate == 4: return int((4*(n**3)) / 5)
    if rate == 5: return int((5*(n**3)) / 4)
    return n ** 3

def calc_level_from_exp(rate, exp):
    """Derive level from EXP using the species growth rate."""
    for lvl in range(1, 101):
        if exp < exp_at_level(rate, lvl + 1):
            return lvl
    return 100

# ---------------------------------------------------------------------------
# Save file parsing
# ---------------------------------------------------------------------------

def find_active_section(data, target_id):
    """Find the section with target_id that has the highest save index."""
    best = None
    for i in range(0, len(data), SECTION_SIZE):
        if i + SECTION_SIZE > len(data):
            break
        sec_id = ru16(data, i + 0xFF4)
        if sec_id == target_id:
            save_idx = ru32(data, i + 0xFFC)
            if best is None or save_idx > best[1]:
                best = (i, save_idx)
    return best[0] if best else None

def get_active_pc_sectors(data):
    """Return the highest-index set of all PC sectors."""
    sections = []
    for i in range(0, len(data), SECTION_SIZE):
        if i + SECTION_SIZE > len(data):
            break
        sec_id  = ru16(data, i + 0xFF0 + 4)
        save_idx = ru32(data, i + 0xFF0 + 12)
        if sec_id in ALL_PC_SECTORS:
            sections.append({"id": sec_id, "idx": save_idx, "offset": i})
    if not sections:
        return []
    max_idx = max(s["idx"] for s in sections)
    return sorted([s for s in sections if s["idx"] == max_idx], key=lambda x: x["id"])

# ---------------------------------------------------------------------------
# Party parsing
# ---------------------------------------------------------------------------

def parse_party_mon(raw, db_species, db_items, db_moves, db_abilities,
                    db_abilities_meta, db_gender, db_move_pp, db_growth_rates):
    """CFRU/Unbound party mons use the same compact format as PC mons but with
    a 4-byte longer OT_NAME field, shifting species/moves/EVs/IVs offsets by +4."""
    if len(raw) < MON_SIZE_PARTY:
        return None

    pid        = ru32(raw, OFF_PTY_PID)
    otid       = ru32(raw, OFF_PTY_OTID)
    species_id = ru16(raw, OFF_PTY_SPECIES)
    exp        = ru32(raw, OFF_PTY_EXP)

    if pid == 0 and otid == 0:
        return None
    if species_id == 0 or species_id > 2500:
        return None
    if exp == 0 or exp > 2_000_000:
        return None

    nickname = decode_text(raw[OFF_PTY_NICK:OFF_PTY_NICK + 10], CHARMAP_PC)
    item_id  = ru16(raw, OFF_PTY_ITEM)

    # Moves: plain u16 per slot (party format, NOT bitpacked like PC)
    move_ids = [ru16(raw, OFF_PTY_MOVES + i * 2) for i in range(4)]

    # EVs
    ev_hp  = raw[OFF_PTY_EVS + 0]; ev_atk = raw[OFF_PTY_EVS + 1]
    ev_def = raw[OFF_PTY_EVS + 2]; ev_spe = raw[OFF_PTY_EVS + 3]
    ev_spa = raw[OFF_PTY_EVS + 4]; ev_spd = raw[OFF_PTY_EVS + 5]

    # IVs + HA flag
    iv_packed = ru32(raw, OFF_PTY_IVS)
    ivs = unpack_ivs(iv_packed)
    ha_flag = (iv_packed >> 31) & 1

    nature_id = pid % 25
    gender_threshold = db_gender.get(species_id)
    gender = gender_from_pid(pid, gender_threshold)
    shiny  = is_shiny(otid, pid)

    ab_meta = db_abilities_meta.get(species_id, {})
    if ha_flag:
        ab_id   = ab_meta.get("ha")
        ab_slot = "HA"
    else:
        slot_bit = pid & 1
        ab_id   = ab_meta.get("a2") if slot_bit else ab_meta.get("a1")
        ab_slot = "Slot 2" if slot_bit else "Slot 1"
    ability_name = db_abilities.get(ab_id, "") if (ab_id is not None and ab_id != 0) else ""

    species_name = db_species.get(species_id, f"#{species_id}")
    item_name    = db_items.get(item_id, "") if item_id > 0 else ""
    move_names   = [db_moves.get(m, f"Move#{m}") if m > 0 else "" for m in move_ids]

    growth_rate = db_growth_rates.get(species_id, 0)
    level = calc_level_from_exp(growth_rate, exp)

    # Stats stored in the party-only bytes after the compact block
    def read_stat(off):
        if off + 2 <= len(raw):
            v = ru16(raw, off)
            return v if 1 <= v <= 999 else ""
        return ""

    max_hp   = read_stat(OFF_PTY_MAX_HP)
    stat_atk = read_stat(OFF_PTY_ATK)
    stat_def = read_stat(OFF_PTY_DEF)
    stat_spe = read_stat(OFF_PTY_SPE)
    stat_spa = read_stat(OFF_PTY_SPA)
    stat_spd = read_stat(OFF_PTY_SPD)

    return {
        "nickname":     nickname,
        "species":      species_name,
        "species_id":   species_id,
        "level":        level,
        "gender":       gender,
        "nature":       DB_NATURES.get(nature_id, f"Nature#{nature_id}"),
        "nature_id":    nature_id,
        "shiny":        "★" if shiny else "",
        "ability":      ability_name,
        "ability_slot": ab_slot,
        "item":         item_name,
        "hp":           max_hp,
        "atk":          stat_atk,
        "def":          stat_def,
        "spe":          stat_spe,
        "spa":          stat_spa,
        "spd":          stat_spd,
        "ev_hp":  ev_hp,  "ev_atk": ev_atk, "ev_def": ev_def,
        "ev_spe": ev_spe, "ev_spa": ev_spa,  "ev_spd": ev_spd,
        "iv_hp":  ivs["HP"],  "iv_atk": ivs["Atk"], "iv_def": ivs["Def"],
        "iv_spe": ivs["Spe"], "iv_spa": ivs["SpA"],  "iv_spd": ivs["SpD"],
        "move1": move_names[0], "move2": move_names[1],
        "move3": move_names[2], "move4": move_names[3],
        "pid":    f"{pid:08X}",
        "location": "Party",
    }

def read_party(data, db_species, db_items, db_moves, db_abilities,
               db_abilities_meta, db_gender, db_move_pp, db_growth_rates):
    trainer_off = find_active_section(data, TRAINER_SECTION_ID)
    if trainer_off is None:
        print("ERROR: Could not find trainer section (Section ID 1). Is this a valid save?")
        return []

    sec = data[trainer_off:trainer_off + SECTION_SIZE]
    count = ru32(sec, PARTY_COUNT_OFFSET)
    count = min(count, 6)

    party = []
    for i in range(count):
        mon_off = PARTY_OFFSET + i * MON_SIZE_PARTY
        raw = sec[mon_off:mon_off + MON_SIZE_PARTY]
        mon = parse_party_mon(raw, db_species, db_items, db_moves, db_abilities,
                              db_abilities_meta, db_gender, db_move_pp, db_growth_rates)
        if mon:
            mon["location"] = f"Party Slot {i+1}"
            party.append(mon)

    print(f"  Party: {len(party)} Pokemon")
    return party

# ---------------------------------------------------------------------------
# PC parsing
# ---------------------------------------------------------------------------

def parse_pc_mon(raw, box_num, slot_num, db_species, db_items, db_moves,
                 db_abilities, db_abilities_meta, db_gender, db_growth_rates):
    if len(raw) < MON_SIZE_PC:
        return None

    pid      = ru32(raw, OFF_PC_PID)
    otid     = ru32(raw, OFF_PC_OTID)
    species_id = ru16(raw, OFF_PC_SPECIES)
    exp      = ru32(raw, OFF_PC_EXP)

    if species_id == 0 or species_id > 2500:
        return None
    if exp == 0 or exp > 2_000_000:
        return None

    nickname  = decode_text(raw[OFF_PC_NICK:OFF_PC_NICK + 10], CHARMAP_PC)
    item_id   = ru16(raw, OFF_PC_ITEM)

    # Moves: bitpacked 10 bits each starting at 0x27
    packed_moves = 0
    for i in range(5):
        packed_moves |= raw[0x27 + i] << (8 * i)
    move_ids = [
        (packed_moves >>  0) & 0x3FF,
        (packed_moves >> 10) & 0x3FF,
        (packed_moves >> 20) & 0x3FF,
        (packed_moves >> 30) & 0x3FF,
    ]

    # EVs
    ev_hp  = raw[OFF_PC_EVS + 0]; ev_atk = raw[OFF_PC_EVS + 1]
    ev_def = raw[OFF_PC_EVS + 2]; ev_spe = raw[OFF_PC_EVS + 3]
    ev_spa = raw[OFF_PC_EVS + 4]; ev_spd = raw[OFF_PC_EVS + 5]

    # IVs
    iv_packed = ru32(raw, OFF_PC_IVS)
    ivs = unpack_ivs(iv_packed)
    ha_flag = (iv_packed >> 31) & 1

    nature_id = pid % 25
    gender_threshold = db_gender.get(species_id)
    gender = gender_from_pid(pid, gender_threshold)
    shiny  = is_shiny(otid, pid)

    ab_meta = db_abilities_meta.get(species_id, {})
    if ha_flag:
        ab_id   = ab_meta.get("ha")
        ab_slot = "HA"
    else:
        slot_bit = pid & 1
        ab_id   = ab_meta.get("a2") if slot_bit else ab_meta.get("a1")
        ab_slot = "Slot 2" if slot_bit else "Slot 1"
    ability_name = db_abilities.get(ab_id, "") if (ab_id is not None and ab_id != 0) else ""

    species_name = db_species.get(species_id, f"#{species_id}")
    item_name    = db_items.get(item_id, "") if item_id > 0 else ""
    move_names   = [db_moves.get(m, f"Move#{m}") if m > 0 else "" for m in move_ids]

    # Derive level from EXP using species growth rate
    growth_rate = db_growth_rates.get(species_id, 0)  # default Medium Fast
    level = calc_level_from_exp(growth_rate, exp)
    box_label = "Preset Box" if box_num == 26 else f"Box {box_num}"

    return {
        "nickname":    nickname,
        "species":     species_name,
        "species_id":  species_id,
        "level":       level,
        "gender":      gender,
        "nature":      DB_NATURES.get(nature_id, f"Nature#{nature_id}"),
        "nature_id":   nature_id,
        "shiny":       "★" if shiny else "",
        "ability":     ability_name,
        "ability_slot": ab_slot,
        "item":        item_name,
        "hp":  "", "atk": "", "def": "", "spe": "", "spa": "", "spd": "",
        "ev_hp":  ev_hp,  "ev_atk": ev_atk, "ev_def": ev_def,
        "ev_spe": ev_spe, "ev_spa": ev_spa,  "ev_spd": ev_spd,
        "iv_hp":  ivs["HP"],  "iv_atk": ivs["Atk"], "iv_def": ivs["Def"],
        "iv_spe": ivs["Spe"], "iv_spa": ivs["SpA"],  "iv_spd": ivs["SpD"],
        "move1": move_names[0], "move2": move_names[1],
        "move3": move_names[2], "move4": move_names[3],
        "exp":    exp,
        "pid":    f"{pid:08X}",
        "location": f"{box_label} Slot {slot_num}",
    }

def read_pc(data, db_species, db_items, db_moves, db_abilities,
            db_abilities_meta, db_gender, db_move_pp, db_growth_rates):
    sectors = get_active_pc_sectors(data)
    if not sectors:
        print("WARNING: Could not find PC sectors.")
        return []

    pc_mons = []

    # ------------------------------------------------------------------
    # Build stream buffer from sectors 5-12 (boxes 1-19, partial)
    # Stream = 8 * 0xFF0 = 32,640 bytes = 562 slots = 18 full boxes +
    # 22 slots of box 19. Slot indices are 0-based.
    # ------------------------------------------------------------------
    stream_secs = sorted([s for s in sectors if s["id"] in POKEMON_STREAM_SECTORS],
                         key=lambda x: x["id"])
    stream = bytearray()
    for sec in stream_secs:
        off = sec["offset"]
        stream += data[off + SECTOR_HEADER_SIZE : off + SECTOR_HEADER_SIZE + SECTOR_PAYLOAD_SIZE]

    def read_stream_slot(box, slot):
        """Read a mon from the stream buffer (boxes 1-19)."""
        idx = (box - 1) * 30 + (slot - 1)
        off = idx * MON_SIZE_PC
        if off + MON_SIZE_PC > len(stream):
            return None
        return stream[off : off + MON_SIZE_PC]

    # ------------------------------------------------------------------
    # Fallback box layouts (from PUSE pc.js FALLBACK_BOX_LAYOUTS)
    # 'absolute': offset is absolute within the full save buffer
    # 'section':  offset is relative to a specific section's start
    #
    # Box 20: absolute 0x1EB0C, slots 1-21
    # Box 21: absolute 0x1F1E8, slots 1-30
    # Box 22: absolute 0x1F8B4, slots 1-30
    # Box 23: section 2 @ 0x0F18 slots 1-4, section 3 @ 0x0010 slots 5-30
    # Box 24: section 3 @ 0x05F4, slots 1-30
    # ------------------------------------------------------------------
    FALLBACK_BOX_LAYOUTS = {
        20: [('absolute', 1, 21, 0x1EB0C)],
        21: [('absolute', 1, 30, 0x1F1E8)],
        22: [('absolute', 1, 30, 0x1F8B4)],
        23: [('section',  2, 1,  4,  0x0F18),
             ('section',  3, 5,  30, 0x0010)],
        24: [('section',  3, 1,  30, 0x05F4)],
    }

    # Resolve active section offsets for sections 2 and 3 (fallback sectors)
    fallback_section_offsets = {}
    SECTION_SIZE = 0x1000
    best = {}
    for i in range(0, len(data), SECTION_SIZE):
        if i + SECTION_SIZE > len(data): break
        sid  = ru16(data, i + 0xFF4)
        sidx = ru32(data, i + 0xFFC)
        if sid in (2, 3):
            if sid not in best or sidx > best[sid][0]:
                best[sid] = (sidx, i)
    for sid, (_, off) in best.items():
        fallback_section_offsets[sid] = off

    def read_fallback_slot(box, slot):
        """Read a mon from fallback box storage (boxes 20-24)."""
        layout = FALLBACK_BOX_LAYOUTS.get(box, [])
        for seg in layout:
            kind = seg[0]
            if kind == 'absolute':
                _, start_slot, end_slot, base_off = seg
                if start_slot <= slot <= end_slot:
                    off = base_off + (slot - start_slot) * MON_SIZE_PC
                    if off + MON_SIZE_PC > len(data):
                        return None
                    return data[off : off + MON_SIZE_PC]
            elif kind == 'section':
                _, sec_id, start_slot, end_slot, rel_off = seg
                if start_slot <= slot <= end_slot:
                    sec_off = fallback_section_offsets.get(sec_id)
                    if sec_off is None:
                        return None
                    off = sec_off + rel_off + (slot - start_slot) * MON_SIZE_PC
                    if off + MON_SIZE_PC > len(data):
                        return None
                    return data[off : off + MON_SIZE_PC]
        return None

    # ------------------------------------------------------------------
    # Read all boxes
    # ------------------------------------------------------------------
    box_count = 0
    for box in range(1, 25):  # boxes 1-24
        slots = 21 if box == 20 else 30  # box 20 only has 21 slots
        for slot in range(1, slots + 1):
            if box <= 19:
                raw = read_stream_slot(box, slot)
            else:
                raw = read_fallback_slot(box, slot)
            if raw is None or all(b == 0 for b in raw):
                continue
            mon = parse_pc_mon(raw, box, slot, db_species, db_items, db_moves,
                               db_abilities, db_abilities_meta, db_gender, db_growth_rates)
            if mon:
                pc_mons.append(mon)
                box_count = box

    # ------------------------------------------------------------------
    # Read Preset Box (Box 26) from sector 0 at offset 0xB0
    # ------------------------------------------------------------------
    preset_secs = [s for s in sectors if s["id"] == PRESET_SECTOR_ID]
    if preset_secs:
        sec0_off = preset_secs[0]["offset"]
        preset_start = sec0_off + OFFSET_PRESET_START
        for slot in range(1, PRESET_CAPACITY + 1):
            off = preset_start + (slot - 1) * MON_SIZE_PC
            if off + MON_SIZE_PC > len(data):
                break
            raw = data[off : off + MON_SIZE_PC]
            if all(b == 0 for b in raw):
                continue
            mon = parse_pc_mon(raw, 26, slot, db_species, db_items, db_moves,
                               db_abilities, db_abilities_meta, db_gender, db_growth_rates)
            if mon:
                pc_mons.append(mon)

    print(f"  PC: {len(pc_mons)} Pokemon across {box_count} box(es) + Preset Box")
    return pc_mons

# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Location",    20),
    ("Nickname",    14),
    ("Species",     16),
    ("Shiny",        6),
    ("Gender",       8),
    ("Level",        7),
    ("Nature",      10),
    ("Ability",     20),
    ("Held Item",   18),
    ("HP",           6), ("Atk",  6), ("Def",  6),
    ("Spe",          6), ("SpA",  6), ("SpD",  6),
    ("HP IV",        7), ("Atk IV", 7), ("Def IV", 7),
    ("Spe IV",       7), ("SpA IV", 7), ("SpD IV", 7),
    ("HP EV",        7), ("Atk EV", 7), ("Def EV", 7),
    ("Spe EV",       7), ("SpA EV", 7), ("SpD EV", 7),
    ("EV Total",    8),
    ("Move 1",      18), ("Move 2", 18), ("Move 3", 18), ("Move 4", 18),
    ("PID",         10),
]

def mon_to_row(mon):
    ev_total = (mon["ev_hp"] + mon["ev_atk"] + mon["ev_def"] +
                mon["ev_spe"] + mon["ev_spa"] + mon["ev_spd"])
    return [
        mon["location"],
        mon["nickname"],
        mon["species"],
        mon["shiny"],
        mon["gender"],
        mon["level"],
        mon["nature"],
        mon["ability"],
        mon["item"],
        mon["hp"],    mon["atk"],    mon["def"],
        mon["spe"],   mon["spa"],    mon["spd"],
        mon["iv_hp"], mon["iv_atk"], mon["iv_def"],
        mon["iv_spe"],mon["iv_spa"], mon["iv_spd"],
        mon["ev_hp"], mon["ev_atk"], mon["ev_def"],
        mon["ev_spe"],mon["ev_spa"], mon["ev_spd"],
        ev_total,
        mon["move1"], mon["move2"], mon["move3"], mon["move4"],
        mon["pid"],
    ]

def make_style():
    header_fill    = PatternFill("solid", fgColor="1F3864")  # dark navy
    party_fill     = PatternFill("solid", fgColor="E8F4FD")  # light blue
    pc_fill        = PatternFill("solid", fgColor="F5F5F5")  # light grey
    shiny_fill     = PatternFill("solid", fgColor="FFF9C4")  # light yellow
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font      = Font(name="Arial", size=10)
    center         = Alignment(horizontal="center", vertical="center")
    left           = Alignment(horizontal="left",   vertical="center")
    thin           = Side(style="thin", color="CCCCCC")
    border         = Border(left=thin, right=thin, top=thin, bottom=thin)
    return header_fill, party_fill, pc_fill, shiny_fill, header_font, cell_font, center, left, border

def write_sheet(ws, mons, title):
    header_fill, party_fill, pc_fill, shiny_fill, header_font, cell_font, center, left, border = make_style()

    ws.title = title
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (col_name, col_w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, mon in enumerate(mons, 2):
        row_data = mon_to_row(mon)
        is_party = "Party" in mon["location"]
        is_shiny = mon["shiny"] == "★"
        base_fill = party_fill if is_party else pc_fill

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = cell_font
            cell.border    = border
            cell.fill      = shiny_fill if is_shiny else base_fill

            # Alignment: center most columns, left for name-like columns
            col_name = COLUMNS[col_idx - 1][0]
            if col_name in ("Location", "Nickname", "Species", "Nature",
                            "Ability", "Held Item", "Move 1", "Move 2",
                            "Move 3", "Move 4"):
                cell.alignment = left
            else:
                cell.alignment = center

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

def write_xlsx(all_mons, output_path, save_name):
    wb = openpyxl.Workbook()

    party_mons = [m for m in all_mons if "Party" in m["location"]]
    pc_mons    = [m for m in all_mons if "Party" not in m["location"]]

    # All mons sheet
    ws_all = wb.active
    write_sheet(ws_all, all_mons, "All Pokemon")

    # Party sheet
    if party_mons:
        ws_party = wb.create_sheet("Party")
        write_sheet(ws_party, party_mons, "Party")

    # PC sheet
    if pc_mons:
        ws_pc = wb.create_sheet("PC Boxes")
        write_sheet(ws_pc, pc_mons, "PC Boxes")

    # Summary sheet
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Pokemon Unbound Save Reader"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14)
    ws_sum["A2"] = f"Save file: {save_name}"
    ws_sum["A3"] = f"Total Pokemon: {len(all_mons)}"
    ws_sum["A4"] = f"Party: {len(party_mons)}"
    ws_sum["A5"] = f"PC Boxes: {len(pc_mons)}"
    ws_sum["A7"] = "Notes:"
    ws_sum["A7"].font = Font(name="Arial", bold=True)
    ws_sum["A8"] = "• PC Pokemon levels are derived from EXP using the species growth rate (same formula the game uses)."
    ws_sum["A9"] = "• Starred (★) rows are Shiny Pokemon."
    ws_sum["A10"] = "• Stats (HP/Atk/etc.) are only stored for Party Pokemon; PC Pokemon stats are blank."
    ws_sum["A11"] = "• EV Total max is 510; individual stat max is 252."
    ws_sum["A12"] = "• This tool supports Pokemon Unbound and CFRU-based FireRed ROM hacks."
    ws_sum.column_dimensions["A"].width = 70

    wb.move_sheet("Summary", offset=-(len(wb.sheetnames)-1))

    wb.save(output_path)
    print(f"\nSaved: {output_path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        save_path = input("Drag and drop your .sav file here (or type the path): ").strip().strip('"')
        if not save_path:
            print(__doc__)
            return
        output_path = Path(save_path).parent / (Path(save_path).stem + "_pokemon.xlsx")
    else:
        save_path = sys.argv[1]
        output_path = Path(save_path).parent / (Path(save_path).stem + "_pokemon.xlsx")

    if not Path(save_path).exists():
        print(f"ERROR: Save file not found: {save_path}")
        return

    print(f"Reading: {save_path}")
    data = bytearray(Path(save_path).read_bytes())
    print(f"  File size: {len(data):,} bytes")

    data_dir = find_data_dir()
    if data_dir:
        print(f"  Data dir: {data_dir}")
    db_species, db_items, db_moves, db_abilities, db_abilities_meta, db_gender, db_move_pp, db_growth_rates = \
        load_databases(data_dir)

    print("\nParsing save...")
    party = read_party(data, db_species, db_items, db_moves, db_abilities,
                       db_abilities_meta, db_gender, db_move_pp, db_growth_rates)
    pc    = read_pc   (data, db_species, db_items, db_moves, db_abilities,
                       db_abilities_meta, db_gender, db_move_pp, db_growth_rates)

    all_mons = party + pc
    if not all_mons:
        print("\nNo Pokemon found. Make sure this is a valid .sav file.")
        return

    print(f"\nTotal Pokemon found: {len(all_mons)}")
    print("Writing Excel...")
    write_xlsx(all_mons, output_path, Path(save_path).name)
    print("Done!")

if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        pass
    except Exception as e:
        print(f"\nERROR: {e}")
    finally:
        input("\nPress Enter to exit...")