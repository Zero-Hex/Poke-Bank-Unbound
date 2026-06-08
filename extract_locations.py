"""
Extract Pokémon location data from the Unbound Location Guide xlsx
and output as borrius_locations.json keyed by Pokémon name.

Usage:
  python3 extract_locations.py <path_to_xlsx> [output_path]
"""

import sys
import json
import openpyxl
from collections import defaultdict

XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else "Pokémon Unbound Location Guide v2.1.1.1.xlsx"
OUT_PATH  = sys.argv[2] if len(sys.argv) > 2 else "static/borrius_locations.json"

locations = defaultdict(list)

def add(name, location, method="Wild"):
    if not name or not location:
        return
    name = str(name).strip()
    location = str(location).strip()
    if name in ("X", "None", "Special Encounter", "") or not name:
        return
    entry = f"{location} ({method})" if method != "Wild" else location
    if entry not in locations[name]:
        locations[name].append(entry)

def even_cols(row):
    """Extract values from every other column (0, 2, 4...)"""
    return [row[c] if c < len(row) else None for c in range(0, len(row), 2)]


# ---------------------------------------------------------------------------
# Tab 1: Grass & Cave Encounters
# Row 0 = location names (even cols), rows 1+ = Pokémon names (even cols)
# ---------------------------------------------------------------------------
def parse_grass_cave(ws):
    rows = list(ws.iter_rows(values_only=True))
    locs = even_cols(rows[0])  # Row 0 is always the location header
    for row in rows[1:]:
        for col_idx, pokemon in enumerate(even_cols(row)):
            if pokemon and col_idx < len(locs) and locs[col_idx]:
                add(str(pokemon).strip(), str(locs[col_idx]).strip(), "Wild")

print("Parsing Grass & Cave Encounters...")
parse_grass_cave(openpyxl.load_workbook(XLSX_PATH, read_only=True)["Grass & Cave Encounters"])


# ---------------------------------------------------------------------------
# Tab 2: Surfing, Fishing, Rock Smash
# Structure: section header ("Surfing"/"Fishing"/"Rock Smash"), blank,
#   location row, then Pokémon rows.
#   Fishing has no new location row — uses same locations as Surfing.
#   Within fishing: "Old Rod"/"Good Rod"/"Super Rod" rows signal method.
# ---------------------------------------------------------------------------
def parse_water(ws):
    rows = list(ws.iter_rows(values_only=True))
    surf_locs = None
    current_locs = None
    current_method = "Surfing"
    METHOD_HEADERS = {"Surfing", "Old Rod", "Good Rod", "Super Rod", "Rock Smash", "Fishing"}

    for row in rows:
        non_null = [v for v in row if v is not None]
        if not non_null:
            continue

        # Single-value section header
        if len(non_null) == 1 and str(non_null[0]).strip() in METHOD_HEADERS:
            current_method = str(non_null[0]).strip()
            if current_method == "Fishing":
                current_locs = surf_locs  # reuse surfing locations
            continue

        # Rod method row: all non-null values are the same rod type
        unique_vals = set(str(v).strip() for v in non_null if v)
        if len(unique_vals) == 1 and list(unique_vals)[0] in METHOD_HEADERS:
            current_method = list(unique_vals)[0]
            continue

        # Location row for Surfing: contains location-like strings
        # Heuristic: first item looks like a route/city/area name and there are many
        evens = even_cols(row)
        first = str(evens[0]).strip() if evens[0] else ""
        if current_method == "Surfing" and surf_locs is None and len(non_null) >= 5:
            surf_locs = evens
            current_locs = surf_locs
            continue

        # Pokémon data row
        if current_locs:
            for col_idx, pokemon in enumerate(evens):
                if pokemon and col_idx < len(current_locs) and current_locs[col_idx]:
                    name = str(pokemon).strip()
                    if name not in METHOD_HEADERS and name not in ("X", ""):
                        add(name, str(current_locs[col_idx]).strip(), current_method)

print("Parsing Surfing, Fishing, Rock Smash...")
parse_water(openpyxl.load_workbook(XLSX_PATH, read_only=True)["Surfing, Fishing, Rock Smash"])


# ---------------------------------------------------------------------------
# Tab 3: Gift & Static Encounters
# Rows: Method | Location | Pokémon | Requirement
# ---------------------------------------------------------------------------
def parse_gifts(ws):
    METHODS = {"Gift", "Static", "Mission Reward", "Event"}
    for row in ws.iter_rows(values_only=True):
        vals = [v for v in row if v is not None]
        if len(vals) < 3:
            continue
        method = str(vals[0]).strip()
        if method not in METHODS:
            continue
        location = str(vals[1]).strip()
        pokemon_cell = str(vals[2]).strip()
        for pokemon in pokemon_cell.split("/"):
            add(pokemon.strip(), location, method)

print("Parsing Gift & Static Encounters...")
parse_gifts(openpyxl.load_workbook(XLSX_PATH, read_only=True)["Gift & Static Encounters"])


# ---------------------------------------------------------------------------
# Tab 4: Legendaries & Mythicals
# Structure: name row (even cols), then header row ("Requires"/"Location"),
#   then data rows with (requirement, location) pairs per column.
# ---------------------------------------------------------------------------
def parse_legendaries(ws):
    rows = list(ws.iter_rows(values_only=True))
    i = 0
    SKIP = {"(Sub) Legendary + Mythical Pokémon", "Ultra Beasts", "Requires", "Location", None}

    while i < len(rows):
        row = rows[i]
        evens = even_cols(row)
        non_null = [v for v in evens if v and str(v).strip() not in SKIP]

        # Name row: multiple Pokémon names, none of them "Requires"/"Location"
        if len(non_null) >= 3 and "Requires" not in [str(v) for v in evens if v]:
            names = [str(v).strip() if v else None for v in evens]
            # Scan ahead for location data rows
            j = i + 1
            while j < len(rows) and j < i + 10:
                data_row = rows[j]
                data_evens = even_cols(data_row)
                data_odds  = [data_row[c] if c < len(data_row) else None for c in range(1, len(data_row), 2)]
                # Skip header rows
                if any(str(v) in ("Requires", "Location") for v in data_row if v):
                    j += 1
                    continue
                if all(v is None for v in data_row):
                    break
                # Each odd column is a location for the name in the previous even column
                for name_idx, name in enumerate(names):
                    if not name or name in SKIP:
                        continue
                    loc_col_idx = name_idx  # location is in same position in odds list
                    if loc_col_idx < len(data_odds) and data_odds[loc_col_idx]:
                        loc = str(data_odds[loc_col_idx]).strip()
                        if loc and loc not in SKIP:
                            add(name, loc, "Legendary/Mythical")
                j += 1
        i += 1

print("Parsing Legendaries & Mythicals...")
parse_legendaries(openpyxl.load_workbook(XLSX_PATH, read_only=True)["LegendMythical & Ultra Beasts"])


# ---------------------------------------------------------------------------
# Tab 5: Swarm Schedule
# Day | hour1 | hour2 ... — Pokémon appear in swarms on various routes
# ---------------------------------------------------------------------------
def parse_swarms(ws):
    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], (int, float)):
            continue
        for v in row[1:]:
            if v and isinstance(v, str):
                add(v.strip(), "Various routes (timed swarm)", "Swarm")

print("Parsing Swarm Schedule...")
parse_swarms(openpyxl.load_workbook(XLSX_PATH, read_only=True)["Swarm Schedule"])


# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
output = {name: {"locations": locs} for name, locs in sorted(locations.items())}

with open(OUT_PATH, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\nDone. {len(output)} Pokémon written to {OUT_PATH}")

for sample in ["Snorunt", "Pikachu", "Articuno", "Litleo", "Magikarp", "Tentacool", "Binacle"]:
    if sample in output:
        locs = output[sample]['locations']
        print(f"  {sample} ({len(locs)} entries): {locs[:3]}{'...' if len(locs)>3 else ''}")
