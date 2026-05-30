#!/usr/bin/env python3
"""
Pokemon Unbound Evolution Checker
Reads a .sav file and outputs a spreadsheet listing every Pokemon
in your party and PC boxes that still has a pending evolution.

Uses the actual Unbound evolution table from the game source code.

Usage:
    python unbound_evo_checker.py <save_file.sav> [output.xlsx]

Requires the 'data/' folder from PUSE's backend/data/ next to this script.
"""

import struct
import sys
import re
import os
import json
import math
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    input("\nPress Enter to close...")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Save file constants (shared with save reader)
# ---------------------------------------------------------------------------
SECTION_SIZE        = 0x1000
TRAINER_SECTION_ID  = 1
PARTY_OFFSET        = 0x38
PARTY_COUNT_OFFSET  = 0x34
MON_SIZE_PARTY      = 100
POKEMON_STREAM_SECTORS = [5, 6, 7, 8, 9, 10, 11, 12]
PRESET_SECTOR_ID    = 0
ALL_PC_SECTORS      = POKEMON_STREAM_SECTORS + [13, PRESET_SECTOR_ID]
SECTOR_HEADER_SIZE  = 4
SECTOR_PAYLOAD_SIZE = 0xFF0
MON_SIZE_PC         = 58
BOX_SLOT_COUNT      = 30
OFFSET_PRESET_START = 0xB0
PRESET_CAPACITY     = 30
SHINY_THRESHOLD     = 16

# Party offsets
OFF_PTY_PID     = 0x00
OFF_PTY_OTID    = 0x04
OFF_PTY_NICK    = 0x08
OFF_PTY_SPECIES = 0x20
OFF_PTY_EXP     = 0x24
OFF_PTY_MOVES   = 0x2C
OFF_PTY_EVS     = 0x30
OFF_PTY_IVS     = 0x3A
OFF_PTY_LEVEL   = 0x54

# PC offsets
OFF_PC_PID      = 0x00
OFF_PC_OTID     = 0x04
OFF_PC_NICK     = 0x08
OFF_PC_SPECIES  = 0x1C
OFF_PC_EXP      = 0x20
OFF_PC_IVS      = 0x36

DB_NATURES = {
    0:"Hardy",1:"Lonely",2:"Brave",3:"Adamant",4:"Naughty",
    5:"Bold",6:"Docile",7:"Relaxed",8:"Impish",9:"Lax",
    10:"Timid",11:"Hasty",12:"Serious",13:"Jolly",14:"Naive",
    15:"Modest",16:"Mild",17:"Quiet",18:"Bashful",19:"Rash",
    20:"Calm",21:"Gentle",22:"Sassy",23:"Careful",24:"Quirky"
}

CHARMAP_PC = {
    0x00:" ",0xA1:"0",0xA2:"1",0xA3:"2",0xA4:"3",0xA5:"4",0xA6:"5",
    0xA7:"6",0xA8:"7",0xA9:"8",0xAA:"9",0xAB:"!",0xAC:"?",0xAD:".",
    0xAE:"-",0xBB:"A",0xBC:"B",0xBD:"C",0xBE:"D",0xBF:"E",0xC0:"F",
    0xC1:"G",0xC2:"H",0xC3:"I",0xC4:"J",0xC5:"K",0xC6:"L",0xC7:"M",
    0xC8:"N",0xC9:"O",0xCA:"P",0xCB:"Q",0xCC:"R",0xCD:"S",0xCE:"T",
    0xCF:"U",0xD0:"V",0xD1:"W",0xD2:"X",0xD3:"Y",0xD4:"Z",0xD5:"a",
    0xD6:"b",0xD7:"c",0xD8:"d",0xD9:"e",0xDA:"f",0xDB:"g",0xDC:"h",
    0xDD:"i",0xDE:"j",0xDF:"k",0xE0:"l",0xE1:"m",0xE2:"n",0xE3:"o",
    0xE4:"p",0xE5:"q",0xE6:"r",0xE7:"s",0xE8:"t",0xE9:"u",0xEA:"v",
    0xEB:"w",0xEC:"x",0xED:"y",0xEE:"z",0xB5:"♂",0xB6:"♀",0xFF:""
}

# ---------------------------------------------------------------------------
# Binary helpers
# ---------------------------------------------------------------------------
def ru8(b, o):  return struct.unpack_from("<B", b, o)[0]
def ru16(b, o): return struct.unpack_from("<H", b, o)[0]
def ru32(b, o): return struct.unpack_from("<I", b, o)[0]

def decode_text(data):
    s = ""
    for byte in data:
        if byte == 0xFF: break
        s += CHARMAP_PC.get(byte, "?")
    return s.strip()

def is_shiny(otid, pid):
    tid = otid & 0xFFFF
    sid = (otid >> 16) & 0xFFFF
    return (tid ^ sid ^ (pid & 0xFFFF) ^ ((pid >> 16) & 0xFFFF)) < SHINY_THRESHOLD

def exp_at_level(rate, n):
    if n <= 1: return 0
    if n > 100: n = 100
    if rate == 0: return n ** 3
    if rate == 1:
        if n <= 50:  return int((n**3*(100-n))/50)
        if n <= 68:  return int((n**3*(150-n))/100)
        if n <= 98:  return int((n**3*((1911-10*n)/3))/500)
        return int((n**3*(160-n))/100)
    if rate == 2:
        if n <= 15:  return int(n**3*((math.floor((n+1)/3)+24)/50))
        if n <= 36:  return int(n**3*((n+14)/50))
        return int(n**3*((math.floor(n/2)+32)/50))
    if rate == 3: return int(1.2*(n**3)-15*(n**2)+100*n-140)
    if rate == 4: return int((4*(n**3))/5)
    if rate == 5: return int((5*(n**3))/4)
    return n**3

def calc_level(rate, exp):
    for lvl in range(1, 101):
        if exp < exp_at_level(rate, lvl+1): return lvl
    return 100

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def find_data_dir():
    candidates = [
        Path(__file__).parent / "data",
        Path.cwd() / "data",
        Path(__file__).parent / "backend" / "data",
    ]
    for p in candidates:
        if (p / "pokemon.txt").exists():
            return p
    return None

def load_id_name(path):
    out = {}
    if not path or not Path(path).exists(): return out
    for line in Path(path).read_text(encoding="utf-8", errors="ignore").splitlines():
        if ":" in line:
            l, r = line.split(":", 1)
            if l.strip().isdigit():
                out[int(l.strip())] = r.strip()
    return out

def load_json(path):
    try: return json.loads(Path(path).read_text(encoding="utf-8"))
    except: return {}

def load_databases(data_dir):
    db_species  = load_id_name(data_dir / "pokemon.txt")
    db_items    = load_id_name(data_dir / "items.txt")

    raw_growth = load_json(data_dir / "species_growth_rates.json")
    db_growth = {}
    for sid, meta in raw_growth.items():
        if str(sid).isdigit() and isinstance(meta, dict):
            rate = meta.get("growth_rate")
            if rate is not None and 0 <= int(rate) <= 5:
                db_growth[int(sid)] = int(rate)

    raw_gender = load_json(data_dir / "species_identity_meta.json")
    db_gender = {}
    for sid, meta in raw_gender.items():
        if str(sid).isdigit() and isinstance(meta, dict):
            t = meta.get("gender_threshold")
            if t is not None:
                db_gender[int(sid)] = int(t) & 0xFF

    db_moves = load_id_name(data_dir / "moves.txt")
    print(f"  Loaded: {len(db_species)} species, {len(db_items)} items, {len(db_moves)} moves, {len(db_growth)} growth rates")
    return db_species, db_items, db_moves, db_growth, db_gender

# ---------------------------------------------------------------------------
# Evolution table parser
# ---------------------------------------------------------------------------

# Human-readable descriptions for each method
EVO_METHOD_DESC = {
    "EVO_LEVEL":                    "Level {param}",
    "EVO_LEVEL_DAY":                "Level {param} (daytime)",
    "EVO_LEVEL_NIGHT":              "Level {param} (nighttime)",
    "EVO_LEVEL_ATK_GT_DEF":         "Level {param} (ATK > DEF)",
    "EVO_LEVEL_ATK_EQ_DEF":         "Level {param} (ATK = DEF)",
    "EVO_LEVEL_ATK_LT_DEF":         "Level {param} (ATK < DEF)",
    "EVO_LEVEL_SILCOON":            "Level {param} (silcoon path)",
    "EVO_LEVEL_CASCOON":            "Level {param} (cascoon path)",
    "EVO_LEVEL_NINJASK":            "Level {param} (ninjask path)",
    "EVO_LEVEL_SHEDINJA":           "Level {param} (shedinja path)",
    "EVO_LEVEL_HOLD_ITEM":          "Level {param} holding {item}",
    "EVO_LEVEL_SPECIFIC_TIME_RANGE":"Level {param} (specific time)",
    "EVO_MALE_LEVEL":               "Level {param} (male only)",
    "EVO_FEMALE_LEVEL":             "Level {param} (female only)",
    "EVO_FRIENDSHIP":               "Friendship (any time)",
    "EVO_FRIENDSHIP_DAY":           "Friendship (daytime)",
    "EVO_FRIENDSHIP_NIGHT":         "Friendship (nighttime)",
    "EVO_ITEM":                     "Use {item}",
    "EVO_ITEM_NIGHT":               "Use {item} (nighttime)",
    "EVO_ITEM_LOCATION":            "Use {item} (specific location)",
    "EVO_ITEM_HOLD_ITEM":           "Use item while holding {item}",
    "EVO_TRADE":                    "Trade",
    "EVO_TRADE_ITEM":               "Trade holding {item}",
    "EVO_HOLD_ITEM_DAY":            "Level up holding {item} (daytime)",
    "EVO_HOLD_ITEM_NIGHT":          "Level up holding {item} (nighttime)",
    "EVO_MOVE":                     "Know move: {param_name}",
    "EVO_MOVE_TYPE":                "Know a {param_name}-type move",
    "EVO_MOVE_MALE":                "Know move (male only)",
    "EVO_MOVE_FEMALE":              "Know move (female only)",
    "EVO_BEAUTY":                   "High Beauty stat",
    "EVO_MAP":                      "Level up at specific location",
    "EVO_RAINY_FOGGY_OW":           "Level up in rain/fog",
    "EVO_TYPE_IN_PARTY":            "Level up with {param_name}-type in party",
    "EVO_OTHER_PARTY_MON":          "Level up with {param_name} in party",
    "EVO_FLAG_SET":                 "Level up near special rock/location",
    "EVO_CRITICAL_HIT":             "Land 3 critical hits in one battle",
    "EVO_NATURE_HIGH":              "Level up (high key nature)",
    "EVO_NATURE_LOW":               "Level up (low key nature)",
    "EVO_DAMAGE_LOCATION":          "Take 49+ damage then walk to tile",
}

TYPE_NAMES = {
    0:"Normal",1:"Fighting",2:"Flying",3:"Poison",4:"Ground",5:"Rock",
    6:"Bug",7:"Ghost",8:"Steel",9:"Fire",10:"Water",11:"Grass",
    12:"Electric",13:"Psychic",14:"Ice",15:"Dragon",16:"Dark",17:"Fairy"
}

def parse_c_defines(path):
    out = {}
    for line in open(path, encoding="utf-8", errors="ignore"):
        m = re.match(r'#define\s+(\w+)\s+(0x[0-9a-fA-F]+|\d+)', line)
        if m:
            out[m.group(1)] = int(m.group(2), 16) if m.group(2).startswith('0x') else int(m.group(2))
    return out

def load_evolution_table(evo_c_path, species_h_path, items_h_path, moves_h_path=None):
    """Parse the Unbound Evolution Table.c into a dict: species_id -> list of evo dicts."""

    species_defs = parse_c_defines(species_h_path)
    item_defs    = parse_c_defines(items_h_path)
    move_defs    = parse_c_defines(moves_h_path) if moves_h_path and Path(moves_h_path).exists() else {}
    item_by_id   = {v: k[5:].replace('_',' ').title() for k,v in item_defs.items() if k.startswith('ITEM_')}
    move_by_id   = {v: k[5:].replace('_',' ').title() for k,v in move_defs.items() if k.startswith('MOVE_')}

    evo_text = open(evo_c_path, encoding="utf-8", errors="ignore").read()

    # Individual evo tuple pattern: {EVO_METHOD, param, SPECIES_TARGET, extra}
    evo_tuple_re = re.compile(r'\{(EVO_\w+|0xFD|0xFE),\s*([^,]+),\s*(SPECIES_\w+),\s*([^}]+)\}')

    evo_data = {}

    # Split by species label
    sections = re.split(r'(?=\[\w+\]\s*=\s*\{)', evo_text)

    for section in sections:
        label_m = re.match(r'\[(\w+)\]\s*=\s*\{', section)
        if not label_m:
            continue
        species_const = label_m.group(1)
        species_id = species_defs.get(species_const)
        if species_id is None:
            continue

        evos = []
        for e in evo_tuple_re.finditer(section):
            method     = e.group(1).strip()
            param_raw  = e.group(2).strip()
            target_raw = e.group(3).strip()
            extra_raw  = e.group(4).strip()

            # Skip Mega/Gigantamax
            if method in ('EVO_MEGA', 'EVO_GIGANTAMAX', '0xFE', '0xFD'):
                continue

            target_id = species_defs.get(target_raw)

            # Resolve param based on method type
            item_methods = {
                'EVO_ITEM', 'EVO_ITEM_NIGHT', 'EVO_ITEM_LOCATION', 'EVO_ITEM_HOLD_ITEM',
                'EVO_TRADE_ITEM', 'EVO_HOLD_ITEM_DAY', 'EVO_HOLD_ITEM_NIGHT',
                'EVO_LEVEL_HOLD_ITEM',
            }
            level_methods = {
                'EVO_LEVEL', 'EVO_LEVEL_DAY', 'EVO_LEVEL_NIGHT', 'EVO_LEVEL_ATK_GT_DEF',
                'EVO_LEVEL_ATK_EQ_DEF', 'EVO_LEVEL_ATK_LT_DEF', 'EVO_LEVEL_SILCOON',
                'EVO_LEVEL_CASCOON', 'EVO_LEVEL_NINJASK', 'EVO_LEVEL_SHEDINJA',
                'EVO_MALE_LEVEL', 'EVO_FEMALE_LEVEL', 'EVO_LEVEL_SPECIFIC_TIME_RANGE',
            }

            param_int = None
            param_name = param_raw
            item_name = ""

            if method in item_methods:
                item_id = item_defs.get(param_raw)
                if item_id is not None:
                    param_int = item_id
                    item_name = item_by_id.get(item_id, param_raw.replace('ITEM_','').replace('_',' ').title())
                    param_name = item_name
            elif method in level_methods or method == 'EVO_LEVEL_HOLD_ITEM':
                try:
                    param_int = int(param_raw, 0)
                    param_name = str(param_int)
                except:
                    param_int = 0
                # For LEVEL_HOLD_ITEM, extra is the item
                if method == 'EVO_LEVEL_HOLD_ITEM':
                    item_id = item_defs.get(extra_raw)
                    item_name = item_by_id.get(item_id, extra_raw.replace('ITEM_','').replace('_',' ').title()) if item_id else extra_raw
            elif method == 'EVO_MOVE_TYPE':
                type_id = species_defs.get(param_raw) or 0
                # TYPE_ constants aren't in species.h, use fallback
                try: type_id = int(param_raw, 0)
                except: type_id = 0
                param_name = TYPE_NAMES.get(type_id, param_raw.replace('TYPE_','').title())
            elif method in ('EVO_FRIENDSHIP', 'EVO_FRIENDSHIP_DAY', 'EVO_FRIENDSHIP_NIGHT'):
                param_int = 0
                param_name = ""
            elif method == 'EVO_TRADE':
                param_int = 0
                param_name = ""
            elif method in ('EVO_MOVE', 'EVO_MOVE_MALE', 'EVO_MOVE_FEMALE'):
                move_id = move_defs.get(param_raw)
                if move_id is not None:
                    param_int = move_id
                    param_name = move_by_id.get(move_id, param_raw.replace('MOVE_','').replace('_',' ').title())
                else:
                    try: param_int = int(param_raw, 0)
                    except: param_int = 0
            else:
                try: param_int = int(param_raw, 0)
                except: param_int = 0

            evos.append({
                'method':     method,
                'param':      param_int if param_int is not None else 0,
                'param_name': param_name,
                'item_name':  item_name,
                'target_id':  target_id,
                'target_raw': target_raw,
                'extra_raw':  extra_raw,
            })

        if evos:
            evo_data[species_id] = evos

    print(f"  Evolution table: {len(evo_data)} species with pending evolutions loaded")
    return evo_data, item_by_id, species_defs

def describe_evolution(evo, db_species, db_moves):
    """Return a human-readable string describing one evolution path."""
    method = evo['method']
    param  = evo['param']
    param_name = evo['param_name']
    item   = evo['item_name']
    target = db_species.get(evo['target_id'], evo['target_raw'].replace('SPECIES_','').replace('_',' ').title())

    # For move-based evolutions, look up the move name
    if method in ('EVO_MOVE', 'EVO_MOVE_MALE', 'EVO_MOVE_FEMALE') and param > 0:
        move_name = db_moves.get(param, param_name.replace('MOVE_','').replace('_',' ').title())
        param_name = move_name

    template = EVO_METHOD_DESC.get(method, method)
    desc = template.format(param=param, item=item, param_name=param_name)

    return target, desc

# ---------------------------------------------------------------------------
# Save parsing (minimal — just species, level, nickname, location, gender, shiny)
# ---------------------------------------------------------------------------

def find_active_section(data, target_id):
    best = None
    for i in range(0, len(data), SECTION_SIZE):
        if i + SECTION_SIZE > len(data): break
        sec_id   = ru16(data, i + 0xFF4)
        save_idx = ru32(data, i + 0xFFC)
        if sec_id == target_id:
            if best is None or save_idx > best[1]:
                best = (i, save_idx)
    return best[0] if best else None

def get_active_pc_sectors(data):
    sections = []
    for i in range(0, len(data), SECTION_SIZE):
        if i + SECTION_SIZE > len(data): break
        sec_id   = ru16(data, i + 0xFF0 + 4)
        save_idx = ru32(data, i + 0xFF0 + 12)
        if sec_id in ALL_PC_SECTORS:
            sections.append({"id": sec_id, "idx": save_idx, "offset": i})
    if not sections: return []
    max_idx = max(s["idx"] for s in sections)
    return sorted([s for s in sections if s["idx"] == max_idx], key=lambda x: x["id"])

def gender_from_pid(pid, threshold):
    if threshold is None:      return "?"
    if threshold == 255:       return "Genderless"
    if threshold == 0:         return "Male"
    if threshold == 254:       return "Female"
    return "Female" if (pid & 0xFF) < threshold else "Male"

def read_party(data, db_species, db_growth, db_gender):
    sec_off = find_active_section(data, TRAINER_SECTION_ID)
    if sec_off is None: return []
    sec = data[sec_off:sec_off + SECTION_SIZE]
    count = min(ru32(sec, PARTY_COUNT_OFFSET), 6)
    party = []
    for i in range(count):
        off = PARTY_OFFSET + i * MON_SIZE_PARTY
        raw = sec[off:off + MON_SIZE_PARTY]
        pid  = ru32(raw, OFF_PTY_PID)
        otid = ru32(raw, OFF_PTY_OTID)
        sid  = ru16(raw, OFF_PTY_SPECIES)
        exp  = ru32(raw, OFF_PTY_EXP)
        if sid == 0 or sid > 2500 or exp == 0 or exp > 2_000_000: continue
        nick  = decode_text(raw[OFF_PTY_NICK:OFF_PTY_NICK+10])
        level = ru8(raw, OFF_PTY_LEVEL)
        gender = gender_from_pid(pid, db_gender.get(sid))
        party.append({
            "species_id": sid,
            "species":    db_species.get(sid, f"#{sid}"),
            "nickname":   nick,
            "level":      level,
            "gender":     gender,
            "shiny":      "★" if is_shiny(otid, pid) else "",
            "location":   f"Party Slot {i+1}",
        })
    return party

def read_pc(data, db_species, db_growth, db_gender):
    sectors = get_active_pc_sectors(data)
    if not sectors: return []

    # ------------------------------------------------------------------
    # Build stream buffer (sectors 5-12, boxes 1-19)
    # ------------------------------------------------------------------
    stream_secs = sorted([s for s in sectors if s["id"] in POKEMON_STREAM_SECTORS], key=lambda x: x["id"])
    stream = bytearray()
    for sec in stream_secs:
        off = sec["offset"]
        stream += data[off + SECTOR_HEADER_SIZE : off + SECTOR_HEADER_SIZE + SECTOR_PAYLOAD_SIZE]

    def read_stream_slot(box, slot):
        off = ((box - 1) * 30 + (slot - 1)) * MON_SIZE_PC
        if off + MON_SIZE_PC > len(stream): return None
        return stream[off : off + MON_SIZE_PC]

    # ------------------------------------------------------------------
    # Fallback box layouts (PUSE pc.js FALLBACK_BOX_LAYOUTS)
    # Box 20: absolute 0x1EB0C, slots 1-21
    # Box 21: absolute 0x1F1E8, slots 1-30
    # Box 22: absolute 0x1F8B4, slots 1-30
    # Box 23: section 2 @ 0x0F18 slots 1-4, section 3 @ 0x0010 slots 5-30
    # Box 24: section 3 @ 0x05F4, slots 1-30
    # ------------------------------------------------------------------
    FALLBACK_BOX_LAYOUTS = {
        20: [('absolute', 1, 21, 0x1EB0C)],
        21: [('absolute', 1, 30, 0x1F1E8)],
        22: [('absolute', 1, 30, 0x1F8B4)],
        23: [('section',  2, 1,  4,  0x0F18),
             ('section',  3, 5,  30, 0x0010)],
        24: [('section',  3, 1,  30, 0x05F4)],
    }

    SECTION_SIZE = 0x1000
    fb_secs = {}
    best = {}
    for i in range(0, len(data), SECTION_SIZE):
        if i + SECTION_SIZE > len(data): break
        sid  = ru16(data, i + 0xFF4)
        sidx = ru32(data, i + 0xFFC)
        if sid in (2, 3):
            if sid not in best or sidx > best[sid][0]:
                best[sid] = (sidx, i)
    for sid, (_, off) in best.items():
        fb_secs[sid] = off

    def read_fallback_slot(box, slot):
        for seg in FALLBACK_BOX_LAYOUTS.get(box, []):
            if seg[0] == 'absolute':
                _, s, e, base = seg
                if s <= slot <= e:
                    off = base + (slot - s) * MON_SIZE_PC
                    return data[off : off + MON_SIZE_PC] if off + MON_SIZE_PC <= len(data) else None
            elif seg[0] == 'section':
                _, sec_id, s, e, rel = seg
                if s <= slot <= e:
                    sec_off = fb_secs.get(sec_id)
                    if sec_off is None: return None
                    off = sec_off + rel + (slot - s) * MON_SIZE_PC
                    return data[off : off + MON_SIZE_PC] if off + MON_SIZE_PC <= len(data) else None
        return None

    # ------------------------------------------------------------------
    # Read all boxes
    # ------------------------------------------------------------------
    mons = []
    for box in range(1, 25):
        slots = 21 if box == 20 else 30
        for slot in range(1, slots + 1):
            raw = read_stream_slot(box, slot) if box <= 19 else read_fallback_slot(box, slot)
            if raw is None or all(b == 0 for b in raw): continue
            pid  = ru32(raw, OFF_PC_PID)
            otid = ru32(raw, OFF_PC_OTID)
            sid  = ru16(raw, OFF_PC_SPECIES)
            exp  = ru32(raw, OFF_PC_EXP)
            if sid == 0 or sid > 2500 or exp == 0 or exp > 2_000_000: continue
            nick  = decode_text(raw[OFF_PC_NICK:OFF_PC_NICK+10])
            rate  = db_growth.get(sid, 0)
            level = calc_level(rate, exp)
            gender = gender_from_pid(pid, db_gender.get(sid))
            mons.append({
                "species_id": sid,
                "species":    db_species.get(sid, f"#{sid}"),
                "nickname":   nick,
                "level":      level,
                "gender":     gender,
                "shiny":      "★" if is_shiny(otid, pid) else "",
                "location":   f"Box {box} Slot {slot}",
            })

    # ------------------------------------------------------------------
    # Preset box (sector 0 at offset 0xB0)
    # ------------------------------------------------------------------
    preset_secs = [s for s in sectors if s["id"] == PRESET_SECTOR_ID]
    if preset_secs:
        sec0_off = preset_secs[0]["offset"]
        for slot in range(1, PRESET_CAPACITY + 1):
            off = sec0_off + OFFSET_PRESET_START + (slot - 1) * MON_SIZE_PC
            if off + MON_SIZE_PC > len(data): break
            raw = data[off:off + MON_SIZE_PC]
            if all(b == 0 for b in raw): continue
            pid  = ru32(raw, OFF_PC_PID)
            otid = ru32(raw, OFF_PC_OTID)
            sid  = ru16(raw, OFF_PC_SPECIES)
            exp  = ru32(raw, OFF_PC_EXP)
            if sid == 0 or sid > 2500 or exp == 0 or exp > 2_000_000: continue
            nick  = decode_text(raw[OFF_PC_NICK:OFF_PC_NICK+10])
            rate  = db_growth.get(sid, 0)
            level = calc_level(rate, exp)
            gender = gender_from_pid(pid, db_gender.get(sid))
            mons.append({
                "species_id": sid,
                "species":    db_species.get(sid, f"#{sid}"),
                "nickname":   nick,
                "level":      level,
                "gender":     gender,
                "shiny":      "★" if is_shiny(otid, pid) else "",
                "location":   f"Preset Box Slot {slot}",
            })

    return mons

# ---------------------------------------------------------------------------
# Build evolution rows
# ---------------------------------------------------------------------------

def build_evo_rows(all_mons, evo_table, db_species, db_moves):
    """For each mon, look up its evolutions and build output rows."""
    rows = []
    for mon in all_mons:
        sid   = mon["species_id"]
        level = mon["level"]
        evos  = evo_table.get(sid, [])
        if not evos:
            continue

        for evo in evos:
            target_name, how = describe_evolution(evo, db_species, db_moves)
            method = evo['method']

            # "Ready now?" heuristic
            ready = ""
            if method == 'EVO_LEVEL' and isinstance(evo['param'], int) and evo['param'] > 0:
                diff = evo['param'] - level
                ready = "✓ Ready" if level >= evo['param'] else f"Need {diff} more level{'s' if diff != 1 else ''}"
            elif method in ('EVO_FRIENDSHIP', 'EVO_FRIENDSHIP_DAY', 'EVO_FRIENDSHIP_NIGHT'):
                ready = "Check friendship"
            elif method == 'EVO_TRADE':
                ready = "Need to trade"
            elif method == 'EVO_TRADE_ITEM':
                ready = f"Trade holding {evo['item_name']}"
            elif 'ITEM' in method:
                ready = f"Use {evo['item_name']}"
            elif 'LEVEL' in method and isinstance(evo['param'], int) and evo['param'] > 0:
                diff = evo['param'] - level
                ready = "✓ Ready" if level >= evo['param'] else f"Need {diff} more level{'s' if diff != 1 else ''}"

            rows.append({
                "location":    mon["location"],
                "nickname":    mon["nickname"],
                "species":     mon["species"],
                "shiny":       mon["shiny"],
                "gender":      mon["gender"],
                "level":       level,
                "evolves_to":  target_name,
                "how":         how,
                "ready":       ready,
            })

    return rows

# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

COLUMNS = [
    ("Location",    20),
    ("Nickname",    14),
    ("Species",     16),
    ("Shiny",        6),
    ("Gender",       8),
    ("Level",        7),
    ("Evolves To",  16),
    ("How",         36),
    ("Status",      28),
]

def write_xlsx(rows, output_path, save_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending Evolutions"
    ws.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F3864")
    ready_fill  = PatternFill("solid", fgColor="C6EFCE")  # green
    level_fill  = PatternFill("solid", fgColor="FFEB9C")  # yellow
    item_fill   = PatternFill("solid", fgColor="BDD7EE")  # blue
    trade_fill  = PatternFill("solid", fgColor="F4CCCC")  # red
    shiny_fill  = PatternFill("solid", fgColor="FFF9C4")  # gold
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell_font   = Font(name="Arial", size=10)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers
    for col_idx, (col_name, col_w) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_w
    ws.row_dimensions[1].height = 20

    # Data
    for row_idx, row in enumerate(rows, 2):
        values = [
            row["location"], row["nickname"], row["species"], row["shiny"],
            row["gender"], row["level"], row["evolves_to"], row["how"], row["ready"],
        ]
        status = row["ready"]
        is_shiny_mon = row["shiny"] == "★"

        # Pick row colour by status
        if is_shiny_mon:
            base_fill = shiny_fill
        elif "✓ Ready" in status:
            base_fill = ready_fill
        elif "Need" in status:
            base_fill = level_fill
        elif "Trade" in status or "trade" in status:
            base_fill = trade_fill
        else:
            base_fill = item_fill

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font   = cell_font
            cell.border = border
            cell.fill   = base_fill
            col_name = COLUMNS[col_idx - 1][0]
            if col_name in ("Location", "Nickname", "Species", "Evolves To", "How", "Status"):
                cell.alignment = left
            else:
                cell.alignment = center

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Pokemon Unbound Evolution Checker"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)
    ws2["A2"] = f"Save file: {save_name}"
    ws2["A3"] = f"Pokemon with pending evolutions: {len(rows)}"

    ready   = sum(1 for r in rows if "✓ Ready" in r["ready"])
    level   = sum(1 for r in rows if "Need" in r["ready"])
    item_ct = sum(1 for r in rows if "Use " in r["ready"])
    trade   = sum(1 for r in rows if "trade" in r["ready"].lower() or "Trade" in r["ready"])

    ws2["A5"] = "Breakdown:"
    ws2["A5"].font = Font(name="Arial", bold=True)
    ws2["A6"]  = f"  ✓ Ready to evolve now:  {ready}"
    ws2["A7"]  = f"  Need more levels:        {level}"
    ws2["A8"]  = f"  Need an item:            {item_ct}"
    ws2["A9"]  = f"  Need to trade:           {trade}"
    ws2["A10"] = f"  Other condition:         {len(rows) - ready - level - item_ct - trade}"
    ws2["A12"] = "Colour key:"
    ws2["A12"].font = Font(name="Arial", bold=True)

    for row_n, (label, fill) in enumerate([
        ("✓ Ready to evolve now", ready_fill),
        ("Needs more levels",     level_fill),
        ("Needs an item / trade", item_fill),
        ("Needs a trade",         trade_fill),
        ("Shiny",                 shiny_fill),
    ], 13):
        cell = ws2.cell(row=row_n, column=1, value=f"  {label}")
        cell.fill = fill
        cell.font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 40
    wb.move_sheet("Summary", offset=-(len(wb.sheetnames)-1))

    wb.save(output_path)
    print(f"Saved: {output_path}")

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def find_evo_files():
    """Look for the evolution C files next to the script or in common locations."""
    base_names = ["Evolution Table.c", "Evolution_Table.c"]
    candidates = [Path(__file__).parent, Path.cwd()]
    for base in candidates:
        for name in base_names:
            p = base / name
            if p.exists():
                return p
    return None

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("Drag and drop your .sav file onto this script, or run:")
        print("  python unbound_evo_checker.py MySave.sav")
        input("\nPress Enter to close...")
        sys.exit(1)

    save_path   = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else Path(save_path).stem + "_evolutions.xlsx"

    if not Path(save_path).exists():
        print(f"ERROR: Save file not found: {save_path}")
        input("\nPress Enter to close...")
        sys.exit(1)

    print(f"Reading: {save_path}")
    data = bytearray(Path(save_path).read_bytes())

    # Data dir
    data_dir = None
    for p in [Path(__file__).parent / "data", Path.cwd() / "data"]:
        if (p / "pokemon.txt").exists():
            data_dir = p
            break
    if data_dir is None:
        print("ERROR: Could not find data/ directory with pokemon.txt")
        input("\nPress Enter to close...")
        sys.exit(1)
    print(f"  Data dir: {data_dir}")

    # Evolution files — look next to the script
    script_dir = Path(__file__).parent
    evo_c      = script_dir / "Evolution Table.c"
    species_h  = script_dir / "species.h"
    items_h    = script_dir / "items.h"
    moves_h    = script_dir / "moves.h"

    missing = [str(p) for p in [evo_c, species_h, items_h, moves_h] if not p.exists()]
    if missing:
        print("\nERROR: Missing evolution data files. Place these next to the script:")
        for f in missing:
            print(f"  {f}")
        print("\nDownload from:")
        print("  https://raw.githubusercontent.com/Skeli789/Dynamic-Pokemon-Expansion/Unbound/src/Evolution%20Table.c")
        print("  https://raw.githubusercontent.com/Skeli789/Dynamic-Pokemon-Expansion/Unbound/include/species.h")
        print("  https://raw.githubusercontent.com/Skeli789/Dynamic-Pokemon-Expansion/Unbound/include/items.h")
        print("  https://raw.githubusercontent.com/Skeli789/Dynamic-Pokemon-Expansion/Unbound/include/moves.h")
        input("\nPress Enter to close...")
        sys.exit(1)

    print("\nLoading data...")
    db_species, db_items, db_moves, db_growth, db_gender = load_databases(data_dir)
    evo_table, item_by_id, species_defs = load_evolution_table(str(evo_c), str(species_h), str(items_h), str(moves_h))

    print("\nParsing save...")
    party = read_party(data, db_species, db_growth, db_gender)
    pc    = read_pc   (data, db_species, db_growth, db_gender)
    all_mons = party + pc
    print(f"  Found: {len(party)} party, {len(pc)} PC = {len(all_mons)} total")

    print("\nChecking evolutions...")
    rows = build_evo_rows(all_mons, evo_table, db_species, db_moves)
    print(f"  {len(rows)} pending evolution(s) found")

    if not rows:
        print("\nNo pending evolutions found — everyone is fully evolved!")
        input("\nPress Enter to close...")
        sys.exit(0)

    write_xlsx(rows, output_path, Path(save_path).name)
    print("Done!")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print("\n--- ERROR ---")
        traceback.print_exc()
    finally:
        input("\nPress Enter to close...")